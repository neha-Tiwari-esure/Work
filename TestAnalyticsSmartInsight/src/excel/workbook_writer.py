from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl.styles import PatternFill

from analysis.failure_analyzer import REASON_LABELS, group_failure_details_by_message


def write_workbook(rows: Iterable[Mapping], output_path: str, best_run=None) -> None:
    data = list(rows)
    df = pd.DataFrame(data)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Runs_Raw", index=False)

        if not df.empty and best_run is not None:
            ws = writer.sheets["Runs_Raw"]
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            best_idx = df["pass_rate"].idxmax()
            excel_row = best_idx + 2
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = green_fill

        if not df.empty:
            summary = pd.DataFrame(
                {
                    "metric": ["run_count", "total_tests", "passed", "failed", "not_analysed"],
                    "value": [
                        len(df),
                        df["total_tests"].fillna(0).sum(),
                        df["passed"].fillna(0).sum(),
                        df["failed"].fillna(0).sum(),
                        df["not_analysed"].fillna(0).sum(),
                    ],
                }
            )
            summary.to_excel(writer, sheet_name="Sprint_Summary", index=False)


def write_failure_analysis_workbook(
    grouped_rows: Mapping[str, Iterable[Mapping]],
    output_path: str,
    summary_counts: Mapping[str, Mapping[str, int]] | None = None,
) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    ordered_reasons = [reason for reason in ("unknown", "auto") if grouped_rows.get(reason)]
    ordered_reasons.extend(reason for reason in grouped_rows if reason not in ordered_reasons and grouped_rows.get(reason))
    grouped_message_rows = group_failure_details_by_message(grouped_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_rows = []

        for reason in ordered_reasons:
            rows = list(grouped_rows.get(reason) or [])
            if not rows:
                continue

            df = pd.DataFrame(rows)
            sheet_name = REASON_LABELS.get(reason, reason.title().replace("_", ""))[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            grouped_df = pd.DataFrame(grouped_message_rows.get(reason) or [])
            if not grouped_df.empty:
                group_sheet_name = f"{sheet_name}_Groups"[:31]
                grouped_df.to_excel(writer, sheet_name=group_sheet_name, index=False)

            reason_summary = dict((summary_counts or {}).get(reason) or {})
            summary_rows.append(
                {
                    "failure_reason": reason,
                    "raw_failure_instances": int(reason_summary.get("raw_failure_instances", 0)),
                    "filtered_unique_test_names": int(reason_summary.get("unique_test_names", df["testName"].nunique() if "testName" in df else 0)),
                    "message_groups": len(grouped_df),
                }
            )

        if summary_rows:
            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
