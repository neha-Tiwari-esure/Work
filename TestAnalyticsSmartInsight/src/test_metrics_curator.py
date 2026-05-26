from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from client.analytics_api import client_from_env
from main import _resolve_sprint_window
from transform.normalize_runs import parse_datetime_value
import yaml


TEMPLATE_CSV = Path("data/Final/ClaimsCoreAutomationMetricsSprintMetrics2026.csv")
OUTPUT_XLSX = Path("data/Final/ClaimsCoreAutomationMetricsSprintMetrics2026.xlsx")
CONFIG_PATH = Path("config/settings.yaml")
PAGE_SIZE = 20
MAX_PAGES = 80

SECTION_CONFIG = {
    "MOTOR": {
        "run_name": "REGRESSION_CLAIMS_MOTOR",
        "header_fill": "F3D7C9",
    },
    "HOME": {
        "run_name": "REGRESSION_CLAIMS_HOME",
        "header_fill": "CFE8F7",
    },
    "Environment Dependant Tests": {
        "run_name": "Claims_APP23_TESTS",
        "header_fill": "DDEAD3",
    },
}


@dataclass
class SprintMetric:
    total: int
    passed: int
    failed: int
    pass_percent: str
    automation_time_hours: Any


@dataclass
class CandidateRun:
    run_id: str
    run_name: str
    started: str
    finished: str
    total: int
    passed: int
    failed: int
    pass_rate: float


RUN_CACHE: dict[int, list[CandidateRun]] = {}


def _load_config() -> dict[str, Any]:
    return yaml.safe_load(CONFIG_PATH.read_text())


def _load_csv_rows() -> list[list[Any]]:
    with TEMPLATE_CSV.open(newline="") as handle:
        rows = [row for row in csv.reader(handle)]

    width = max(len(row) for row in rows)
    return [row + [""] * (width - len(row)) for row in rows]


def _load_existing_rows() -> list[list[Any]]:
    if not OUTPUT_XLSX.exists():
        return _load_csv_rows()

    workbook = load_workbook(OUTPUT_XLSX, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    width = max(len(row) for row in rows)
    return [list(row) + [""] * (width - len(row)) for row in rows]


def _current_sprint_number() -> int:
    today = datetime.now()
    if today <= datetime(2026, 1, 27):
        return 1
    return 2 + ((today - datetime(2026, 1, 28)).days // 14)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Populate the sprint metrics workbook from Analytics runs.")
    parser.add_argument(
        "--sprint-number",
        dest="sprint_numbers",
        type=int,
        action="append",
        help="Sprint number to update. Repeat to update multiple sprints.",
    )
    parser.add_argument(
        "--sprint-range",
        nargs=2,
        metavar=("START", "END"),
        type=int,
        help="Inclusive sprint range to update, e.g. --sprint-range 7 9",
    )
    return parser.parse_args()


def _resolve_target_sprints(args: argparse.Namespace) -> list[int]:
    sprint_numbers = list(args.sprint_numbers or [])
    if args.sprint_range:
        start, end = args.sprint_range
        if end < start:
            raise SystemExit("Sprint range end must be greater than or equal to start.")
        sprint_numbers.extend(range(start, end + 1))

    if not sprint_numbers:
        sprint_numbers = [_current_sprint_number()]

    unique_sprints = sorted({number for number in sprint_numbers if number > 0})
    if not unique_sprints:
        raise SystemExit("At least one positive sprint number is required.")
    return unique_sprints


def _format_date_label(sprint_number: int) -> str:
    start, end = _resolve_sprint_window(sprint_number)
    return f"{start.day} {start.strftime('%b')} - {end.day} {end.strftime('%b')}"


def _collect_runs_for_sprint(client, sprint_number: int) -> list[CandidateRun]:
    if sprint_number in RUN_CACHE:
        return RUN_CACHE[sprint_number]

    sprint_start, sprint_end = _resolve_sprint_window(sprint_number)
    sprint_end = parse_datetime_value(sprint_end.date().isoformat(), end_of_day=True) or sprint_end
    matches: list[CandidateRun] = []

    for page_num in range(1, MAX_PAGES + 1):
        try:
            runs = client.get_runs(include_improper_runs=True, page_num=page_num, page_size=PAGE_SIZE)
        except Exception as exc:  # noqa: BLE001
            print(f"Run discovery stopped at page {page_num} for Sprint-{sprint_number}: {exc}")
            break
        if not runs:
            break

        oldest_started = None
        for run in runs:
            started_text = run.get("started") or run.get("dateTime") or run.get("executedAt")
            started = parse_datetime_value(started_text)
            if started is not None:
                oldest_started = started

            if started is None or not (sprint_start <= started <= sprint_end):
                continue

            counts = run.get("counts") if isinstance(run.get("counts"), dict) else {}
            statuses = counts.get("statuses") if isinstance(counts.get("statuses"), dict) else {}
            total = int(counts.get("total") or 0)
            passed = int(statuses.get("passed") or 0)
            failed = int(statuses.get("failed") or 0)
            pass_rate = (passed / total * 100) if total else 0.0

            matches.append(
                CandidateRun(
                    run_id=str(run.get("id") or ""),
                    run_name=str(run.get("name") or ""),
                    started=str(started_text or ""),
                    finished=str(run.get("finished") or ""),
                    total=total,
                    passed=passed,
                    failed=failed,
                    pass_rate=pass_rate,
                )
            )

        if oldest_started is not None and oldest_started < sprint_start:
            break

    RUN_CACHE[sprint_number] = matches
    return matches


def _iter_matching_runs(client, run_name: str, sprint_number: int) -> list[CandidateRun]:
    normalized_name = run_name.strip().lower()
    return [run for run in _collect_runs_for_sprint(client, sprint_number) if run.run_name.strip().lower() == normalized_name]


def _pick_best_run(matches: list[CandidateRun]) -> CandidateRun | None:
    if not matches:
        return None

    return max(matches, key=lambda match: (match.pass_rate, match.passed, match.total, match.started))


def _derive_automation_time(run: CandidateRun | None, fallback_value: Any) -> Any:
    if run is None:
        return fallback_value

    started = parse_datetime_value(run.started)
    finished = parse_datetime_value(run.finished)
    if started is None or finished is None:
        return fallback_value

    elapsed_hours = (finished - started).total_seconds() / 3600
    if elapsed_hours <= 0 or elapsed_hours > 24:
        return fallback_value
    return int(math.ceil(elapsed_hours))


def _build_metric(client, run_name: str, sprint_number: int, fallback_time: Any) -> SprintMetric | None:
    best_run = _pick_best_run(_iter_matching_runs(client, run_name, sprint_number))
    if best_run is None:
        return None

    return SprintMetric(
        total=best_run.total,
        passed=best_run.passed,
        failed=best_run.failed,
        pass_percent=f"{(best_run.passed / best_run.total * 100):.2f}%" if best_run.total else "0.00%",
        automation_time_hours=_derive_automation_time(best_run, fallback_time),
    )


def _find_row_index(rows: list[list[Any]], first_cell: str, start: int = 0, end: int | None = None) -> int | None:
    upper = len(rows) if end is None else end
    target = first_cell.strip().lower()
    for idx in range(start, upper):
        if str(rows[idx][0]).strip().lower() == target:
            return idx
    return None


def _previous_non_empty_value(rows: list[list[Any]], row_index: int, before_column: int) -> Any:
    for col in range(before_column - 1, 0, -1):
        value = rows[row_index][col]
        if value not in (None, ""):
            return value
    return ""


def _write_metric_values(rows: list[list[Any]], section_start: int, next_section: int, sprint_number: int, metric: SprintMetric) -> None:
    column_index = sprint_number
    total_row = _find_row_index(rows, "Automation Total", start=section_start, end=next_section)
    pass_row = _find_row_index(rows, "Automation Pass", start=section_start, end=next_section)
    fail_row = _find_row_index(rows, "Automation Fail", start=section_start, end=next_section)
    pct_row = _find_row_index(rows, "Automation % Pass", start=section_start, end=next_section)
    time_row = _find_row_index(rows, "Automation Time (h)", start=section_start, end=next_section)

    if None in {total_row, pass_row, fail_row, pct_row, time_row}:
        return

    rows[section_start + 1][column_index] = _format_date_label(sprint_number)
    rows[total_row][column_index] = metric.total
    rows[pass_row][column_index] = metric.passed
    rows[fail_row][column_index] = metric.failed
    rows[pct_row][column_index] = metric.pass_percent
    rows[time_row][column_index] = metric.automation_time_hours


def update_metrics_workbook(sprint_numbers: list[int]) -> Path:
    config = _load_config()
    load_dotenv(Path(".env"))
    client = client_from_env(config["base_url"])
    rows = _load_existing_rows()

    section_rows = [idx for idx, row in enumerate(rows) if str(row[0]).strip() in SECTION_CONFIG]
    section_boundaries = section_rows + [len(rows)]

    for sprint_number in sprint_numbers:
        print(f"Curating metrics for Sprint-{sprint_number}...")
        for section_index, section_start in enumerate(section_rows):
            section_name = str(rows[section_start][0]).strip()
            next_section = section_boundaries[section_index + 1]
            run_name = SECTION_CONFIG[section_name]["run_name"]

            rows[section_start + 1][sprint_number] = _format_date_label(sprint_number)

            if not run_name:
                continue

            time_row = _find_row_index(rows, "Automation Time (h)", start=section_start, end=next_section)
            fallback_time = _previous_non_empty_value(rows, time_row, sprint_number) if time_row is not None else ""
            metric = _build_metric(client, run_name, sprint_number, fallback_time)
            if metric is None:
                print(f"No matching run found for {section_name} in Sprint-{sprint_number}.")
                continue

            _write_metric_values(rows, section_start, next_section, sprint_number, metric)

    _write_workbook(rows)
    return OUTPUT_XLSX


def _coerce_cell_value(row_label: str, value: Any, column_index: int) -> Any:
    if column_index == 1 or value in (None, ""):
        return value

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return ""

    if text.lower() == "link":
        return text

    if row_label in {"Automation % Pass", "Automation Split %", "Manual % Pass", "Manual Split %"}:
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return text
        try:
            return float(text) / 100
        except ValueError:
            return text

    numeric_rows = {
        "Automation Total",
        "Automation Pass",
        "Automation Fail",
        "Automation Time (h)",
        "Manual Total Count",
        "Manual Pass Count",
        "Manual Fail Count",
        "Manual Time (h)",
        "Manual Total",
        "Manual Pass",
        "Manual Fail",
    }
    if row_label in numeric_rows:
        try:
            return int(float(text))
        except ValueError:
            return text

    return text


def _extract_metric_value(rows: list[list[Any]], section_name: str, row_label: str, sprint_number: int) -> Any:
    section_start = _find_row_index(rows, section_name)
    if section_start is None:
        return None

    section_rows = [idx for idx, row in enumerate(rows) if str(row[0]).strip() in SECTION_CONFIG]
    next_section = next((idx for idx in section_rows if idx > section_start), len(rows))
    metric_row = _find_row_index(rows, row_label, start=section_start, end=next_section)
    if metric_row is None or sprint_number >= len(rows[metric_row]):
        return None
    return rows[metric_row][sprint_number]


def _to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def _build_quarterly_dashboard_sheet(wb: Workbook, rows: list[list[Any]]) -> None:
    if "QuarterlyDashboard2026" in wb.sheetnames:
        del wb["QuarterlyDashboard2026"]

    ws = wb.create_sheet("QuarterlyDashboard2026")
    quarter_columns = [
        ("End Q4 '25", None),
        ("End Q1 '26", 6),
        ("End Q2 '26", 12),
        ("End Q3 '26", 18),
        ("End Q4 '26", 24),
    ]
    metric_labels = [
        "Total Combined",
        "Automation Total",
        "Automation Pass",
        "Automation Fail",
        "Automation % Pass",
        "Automation Time (h)",
        "Automation Split",
        "Manual Total",
        "Manual Pass",
        "Manual Fail",
        "Manual % Pass",
        "Manual Time (h)",
        "Manual Split",
    ]

    ws.cell(row=1, column=1, value="Text field")
    for col_index, (header, _sprint_number) in enumerate(quarter_columns, start=2):
        ws.cell(row=1, column=col_index, value=header)
    for row_index, label in enumerate(metric_labels, start=2):
        ws.cell(row=row_index, column=1, value=label)

    for col_index, (_header, sprint_number) in enumerate(quarter_columns, start=2):
        if sprint_number is None:
            continue

        motor_auto_total = _to_number(_extract_metric_value(rows, "MOTOR", "Automation Total", sprint_number)) or 0
        home_auto_total = _to_number(_extract_metric_value(rows, "HOME", "Automation Total", sprint_number)) or 0
        env_auto_total = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Automation Total", sprint_number)) or 0
        automation_total = motor_auto_total + home_auto_total + env_auto_total

        motor_auto_pass = _to_number(_extract_metric_value(rows, "MOTOR", "Automation Pass", sprint_number)) or 0
        home_auto_pass = _to_number(_extract_metric_value(rows, "HOME", "Automation Pass", sprint_number)) or 0
        env_auto_pass = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Automation Pass", sprint_number)) or 0
        automation_pass = motor_auto_pass + home_auto_pass + env_auto_pass

        motor_auto_fail = _to_number(_extract_metric_value(rows, "MOTOR", "Automation Fail", sprint_number)) or 0
        home_auto_fail = _to_number(_extract_metric_value(rows, "HOME", "Automation Fail", sprint_number)) or 0
        env_auto_fail = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Automation Fail", sprint_number)) or 0
        automation_fail = motor_auto_fail + home_auto_fail + env_auto_fail

        motor_auto_time = _to_number(_extract_metric_value(rows, "MOTOR", "Automation Time (h)", sprint_number)) or 0
        home_auto_time = _to_number(_extract_metric_value(rows, "HOME", "Automation Time (h)", sprint_number)) or 0
        env_auto_time = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Automation Time (h)", sprint_number)) or 0
        automation_time = motor_auto_time + home_auto_time + env_auto_time

        motor_manual_total = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Total Count", sprint_number))
        if motor_manual_total is None:
            motor_manual_total = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Total", sprint_number)) or 0
        home_manual_total = _to_number(_extract_metric_value(rows, "HOME", "Manual Total Count", sprint_number))
        if home_manual_total is None:
            home_manual_total = _to_number(_extract_metric_value(rows, "HOME", "Manual Total", sprint_number)) or 0
        env_manual_total = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Total Count", sprint_number))
        if env_manual_total is None:
            env_manual_total = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Total", sprint_number)) or 0
        manual_total = motor_manual_total + home_manual_total + env_manual_total

        motor_manual_pass = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Pass Count", sprint_number))
        if motor_manual_pass is None:
            motor_manual_pass = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Pass", sprint_number)) or 0
        home_manual_pass = _to_number(_extract_metric_value(rows, "HOME", "Manual Pass Count", sprint_number))
        if home_manual_pass is None:
            home_manual_pass = _to_number(_extract_metric_value(rows, "HOME", "Manual Pass", sprint_number)) or 0
        env_manual_pass = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Pass Count", sprint_number))
        if env_manual_pass is None:
            env_manual_pass = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Pass", sprint_number)) or 0
        manual_pass = motor_manual_pass + home_manual_pass + env_manual_pass

        motor_manual_fail = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Fail Count", sprint_number))
        if motor_manual_fail is None:
            motor_manual_fail = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Fail", sprint_number)) or 0
        home_manual_fail = _to_number(_extract_metric_value(rows, "HOME", "Manual Fail Count", sprint_number))
        if home_manual_fail is None:
            home_manual_fail = _to_number(_extract_metric_value(rows, "HOME", "Manual Fail", sprint_number)) or 0
        env_manual_fail = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Fail Count", sprint_number))
        if env_manual_fail is None:
            env_manual_fail = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Fail", sprint_number)) or 0
        manual_fail = motor_manual_fail + home_manual_fail + env_manual_fail

        motor_manual_time = _to_number(_extract_metric_value(rows, "MOTOR", "Manual Time (h)", sprint_number)) or 0
        home_manual_time = _to_number(_extract_metric_value(rows, "HOME", "Manual Time (h)", sprint_number)) or 0
        env_manual_time = _to_number(_extract_metric_value(rows, "Environment Dependant Tests", "Manual Time (h)", sprint_number)) or 0
        manual_time = motor_manual_time + home_manual_time + env_manual_time

        total_combined = automation_total + manual_total
        automation_pass_pct = (automation_pass / automation_total) if automation_total else None
        automation_split = (automation_total / total_combined) if total_combined else None
        manual_pass_pct = (manual_pass / manual_total) if manual_total else None
        manual_split = (manual_total / total_combined) if total_combined else None

        values = [
            total_combined if total_combined else None,
            automation_total if automation_total else None,
            automation_pass if automation_pass else None,
            automation_fail if automation_total else None,
            automation_pass_pct,
            automation_time if automation_time else None,
            automation_split,
            manual_total if manual_total else None,
            manual_pass if manual_pass else None,
            manual_fail if manual_total else None,
            manual_pass_pct,
            manual_time if manual_time else None,
            manual_split,
        ]
        for row_index, value in enumerate(values, start=2):
            ws.cell(row=row_index, column=col_index, value=value)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    highlight_fill = PatternFill(fill_type="solid", start_color="FFF200", end_color="FFF200")

    ws.column_dimensions["A"].width = 24
    for col_idx in range(2, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_idx in range(1, 15):
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = left if col_idx == 1 else center
            if row_idx == 1:
                cell.font = Font(bold=True)

    percent_rows = {6, 8, 12, 14}
    for row_idx in percent_rows:
        for col_idx in range(2, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0%"

    ws["C7"].fill = highlight_fill


def _write_workbook(rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SprintMetrics2026"

    for row_index, row in enumerate(rows, start=1):
        row_label = str(row[0]).strip() if row else ""
        for column_index, raw_value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index, value=_coerce_cell_value(row_label, raw_value, column_index))

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    pct_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

    ws.column_dimensions["A"].width = 30
    for col in range(2, 27):
        ws.column_dimensions[get_column_letter(col)].width = 14

    section_rows = [idx + 1 for idx, row in enumerate(rows) if str(row[0]).strip() in SECTION_CONFIG]
    for row_idx, row in enumerate(rows, start=1):
        row_label = str(ws.cell(row=row_idx, column=1).value).strip()
        max_col = len(row)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = left if col_idx == 1 else center
            if row_idx in section_rows and col_idx >= 2:
                cell.fill = header_fill
                cell.font = Font(bold=True)

            if row_label in {"Automation % Pass", "Automation Split %", "Manual % Pass", "Manual Split %"} and col_idx >= 2:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00%" if "Pass" in row_label else "0%"

        if row_label == "Automation % Pass":
            for col_idx in range(1, max_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = pct_fill

    for row_idx in section_rows:
        section_name = str(ws.cell(row=row_idx, column=1).value).strip()
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)
        cell = ws.cell(row=row_idx, column=1)
        section_fill = PatternFill(
            fill_type="solid",
            start_color=SECTION_CONFIG[section_name]["header_fill"],
            end_color=SECTION_CONFIG[section_name]["header_fill"],
        )
        cell.fill = section_fill
        cell.font = Font(bold=True)
        cell.alignment = center
        ws.cell(row=row_idx + 1, column=1).fill = section_fill
        ws.cell(row=row_idx + 1, column=1).border = border

    ws.freeze_panes = "B2"
    _build_quarterly_dashboard_sheet(wb, rows)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    args = _parse_args()
    sprint_numbers = _resolve_target_sprints(args)
    output_path = update_metrics_workbook(sprint_numbers)
    print(f"Updated workbook: {output_path}")
